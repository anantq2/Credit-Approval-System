import os
import openpyxl
from celery import shared_task
from customers.models import Customer
from loans.models import Loan
from django.utils import timezone
from decimal import Decimal


@shared_task
def ingest_customer_data():
    file_path = os.path.join('excel_data', 'customer_data.xlsx')

    if not os.path.exists(file_path):
        print(f"❌ Customer Excel file not found at: {file_path}")
        return "Customer Excel file not found"

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"❌ Failed to load customer Excel file: {e}")
        return f"Failed to load customer Excel file: {e}"

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            first_name, last_name, age, phone_number = row
            Customer.objects.create(
                first_name=first_name,
                last_name=last_name,
                age=int(age),
                phone_number=str(phone_number),
                created_at=timezone.now()
            )
        except Exception as e:
            print(f"❌ Error ingesting customer row {row}: {e}")

    print("✅ Customer data ingestion complete")
    return "Customer data ingestion complete"


@shared_task
def ingest_loan_data():
    file_path = os.path.join('excel_data', 'loan_data.xlsx')

    if not os.path.exists(file_path):
        print(f"❌ Loan Excel file not found at: {file_path}")
        return "Loan Excel file not found"

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
    except Exception as e:
        print(f"❌ Failed to load loan Excel file: {e}")
        return f"Failed to load loan Excel file: {e}"

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            customer_id, loan_amount, tenure, interest_rate, monthly_repayment, emis_paid_on_time, start_date, end_date = row
            customer = Customer.objects.filter(customer_id=customer_id).first()

            if customer:
                Loan.objects.create(
                    customer=customer,
                    loan_amount=Decimal(loan_amount),
                    tenure=int(tenure),
                    interest_rate=Decimal(interest_rate),
                    monthly_repayment=Decimal(monthly_repayment),
                    emis_paid_on_time=int(emis_paid_on_time),
                    start_date=start_date,
                    end_date=end_date,
                    created_at=timezone.now()
                )
            else:
                print(f"⚠️ Customer not found: {customer_id}")
        except Exception as e:
            print(f"❌ Error ingesting loan row {row}: {e}")

    print("✅ Loan data ingestion complete")
    return "Loan data ingestion complete"
